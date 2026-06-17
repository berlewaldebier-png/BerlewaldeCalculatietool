from __future__ import annotations

import csv
import hashlib
import io
import json
import zipfile
import xml.etree.ElementTree as ET
from datetime import UTC, date, datetime, timedelta
from pathlib import Path
from threading import Lock
from typing import Any
from uuid import uuid4

from app.domain import postgres_storage, product_model_storage

_SCHEMA_READY = False
_SCHEMA_LOCK = Lock()


def _now() -> datetime:
    return datetime.now(UTC)


def _text(value: Any) -> str:
    return str(value or "").strip()


def _num(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    text = text.replace("\u20ac", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        # Excel serial date, using the Windows 1900 date system.
        try:
            return date(1899, 12, 30) + timedelta(days=float(value))
        except Exception:
            return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except ValueError:
        return None


def _bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    text = str(value or "").strip().lower()
    return text in {"1", "true", "yes", "ja", "j", "y"}


def _id_for(*parts: Any) -> str:
    raw = "|".join(_text(part).lower() for part in parts)
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()
    return digest[:40]


STOCK_HISTORY_COLUMNS = [
    "Datum",
    "Transactienummer",
    "Artikelnummer",
    "Product",
    "Batchnummer",
    "Bedrijf",
    "Aantal",
]

STOCK_HISTORY_REQUIRED_KEYS = {
    "date",
    "transaction_number",
    "sku_code",
    "product_name",
    "lot_number",
    "company_name",
    "quantity",
}

OPENING_LOT_COLUMNS = [
    "Supplier",
    "LOT nummer",
    "SKU code",
    "Product",
    "Datum",
    "Aantal ref.",
    "Inkoopprijs",
    "Accijns",
    "Inclusief accijns",
]

OPENING_LOT_REQUIRED_KEYS = {
    "supplier",
    "lot_number",
    "sku_code",
    "product_name",
    "source_date",
    "quantity",
    "purchase_price_input",
    "excise_per_unit",
    "purchase_price_includes_excise",
}


def ensure_schema() -> None:
    global _SCHEMA_READY
    if _SCHEMA_READY:
        return
    with _SCHEMA_LOCK:
        if _SCHEMA_READY:
            return
        postgres_storage.ensure_schema()
        with postgres_storage.connect() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS lot_cost_records (
                        id TEXT PRIMARY KEY,
                        source_type TEXT NOT NULL DEFAULT 'purchase_invoice',
                        source_ref TEXT NOT NULL DEFAULT '',
                        supplier TEXT NOT NULL DEFAULT '',
                        lot_number TEXT NOT NULL DEFAULT '',
                        sku_id TEXT NOT NULL DEFAULT '',
                        sku_code TEXT NOT NULL DEFAULT '',
                        product_name TEXT NOT NULL DEFAULT '',
                        source_date DATE,
                        quantity NUMERIC NOT NULL DEFAULT 0,
                        purchase_price_input NUMERIC NOT NULL DEFAULT 0,
                        purchase_price_includes_excise BOOLEAN NOT NULL DEFAULT FALSE,
                        purchase_price_ex_excise NUMERIC NOT NULL DEFAULT 0,
                        excise_per_unit NUMERIC NOT NULL DEFAULT 0,
                        packaging_cost_per_unit NUMERIC NOT NULL DEFAULT 0,
                        other_direct_cost_per_unit NUMERIC NOT NULL DEFAULT 0,
                        payload JSONB NOT NULL DEFAULT '{}'::jsonb,
                        updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                    """
                )
                cur.execute("CREATE INDEX IF NOT EXISTS ix_lot_cost_lot_sku ON lot_cost_records(lot_number, sku_code)")
                cur.execute("CREATE INDEX IF NOT EXISTS ix_lot_cost_sku_id ON lot_cost_records(sku_id)")
                cur.execute("CREATE INDEX IF NOT EXISTS ix_lot_cost_date ON lot_cost_records(source_date)")
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS sales_lot_allocations (
                        id TEXT PRIMARY KEY,
                        import_batch_id TEXT NOT NULL DEFAULT '',
                        source_filename TEXT NOT NULL DEFAULT '',
                        transaction_number TEXT NOT NULL DEFAULT '',
                        sku_code TEXT NOT NULL DEFAULT '',
                        lot_number TEXT NOT NULL DEFAULT '',
                        movement_date DATE,
                        product_name TEXT NOT NULL DEFAULT '',
                        company_name TEXT NOT NULL DEFAULT '',
                        quantity NUMERIC NOT NULL DEFAULT 0,
                        stock_value_per_unit NUMERIC NOT NULL DEFAULT 0,
                        excise_per_unit NUMERIC NOT NULL DEFAULT 0,
                        movement_type TEXT NOT NULL DEFAULT '',
                        movement_reason TEXT NOT NULL DEFAULT '',
                        payload JSONB NOT NULL DEFAULT '{}'::jsonb,
                        updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                    """
                )
                cur.execute("ALTER TABLE sales_lot_allocations ADD COLUMN IF NOT EXISTS import_batch_id TEXT NOT NULL DEFAULT ''")
                cur.execute("ALTER TABLE sales_lot_allocations ADD COLUMN IF NOT EXISTS source_filename TEXT NOT NULL DEFAULT ''")
                cur.execute(
                    "CREATE INDEX IF NOT EXISTS ix_sales_lot_tx_sku ON sales_lot_allocations(transaction_number, sku_code)"
                )
                cur.execute("CREATE INDEX IF NOT EXISTS ix_sales_lot_lot ON sales_lot_allocations(lot_number)")
                cur.execute("CREATE INDEX IF NOT EXISTS ix_sales_lot_import_batch ON sales_lot_allocations(import_batch_id)")
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS lot_alias_mappings (
                        id TEXT PRIMARY KEY,
                        sku_id TEXT NOT NULL DEFAULT '',
                        sku_code TEXT NOT NULL DEFAULT '',
                        douano_lot_number TEXT NOT NULL DEFAULT '',
                        internal_lot_number TEXT NOT NULL DEFAULT '',
                        reason TEXT NOT NULL DEFAULT '',
                        payload JSONB NOT NULL DEFAULT '{}'::jsonb,
                        updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                    """
                )
                cur.execute(
                    """
                    CREATE UNIQUE INDEX IF NOT EXISTS ux_lot_alias_scope
                    ON lot_alias_mappings (
                        LOWER(COALESCE(sku_id, '')),
                        LOWER(COALESCE(sku_code, '')),
                        LOWER(douano_lot_number)
                    )
                    """
                )
                cur.execute("CREATE INDEX IF NOT EXISTS ix_lot_alias_douano_lot ON lot_alias_mappings(douano_lot_number)")
            if not postgres_storage.in_transaction():
                conn.commit()
        _SCHEMA_READY = True


def list_lot_cost_records(limit: int = 2000) -> list[dict[str, Any]]:
    ensure_schema()
    lim = max(1, min(int(limit or 2000), 10000))
    with postgres_storage.connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id, source_type, source_ref, supplier, lot_number, sku_id, sku_code, product_name, source_date,
                       quantity, purchase_price_input, purchase_price_includes_excise, purchase_price_ex_excise,
                       excise_per_unit, packaging_cost_per_unit, other_direct_cost_per_unit, payload, updated_at
                FROM lot_cost_records
                ORDER BY source_date DESC NULLS LAST, updated_at DESC, lot_number, sku_code
                LIMIT %s
                """,
                (lim,),
            )
            rows = cur.fetchall() or []
    out: list[dict[str, Any]] = []
    for row in rows:
        (
            rid,
            source_type,
            source_ref,
            supplier,
            lot_number,
            sku_id,
            sku_code,
            product_name,
            source_date,
            quantity,
            purchase_price_input,
            includes_excise,
            purchase_ex,
            excise,
            packaging,
            other_direct,
            payload,
            updated_at,
        ) = row
        data = payload if isinstance(payload, dict) else {}
        out.append(
            {
                **data,
                "id": _text(rid),
                "source_type": _text(source_type),
                "source_ref": _text(source_ref),
                "supplier": _text(supplier),
                "lot_number": _text(lot_number),
                "sku_id": _text(sku_id),
                "sku_code": _text(sku_code),
                "product_name": _text(product_name),
                "source_date": source_date.isoformat() if source_date else "",
                "quantity": float(quantity or 0),
                "purchase_price_input": float(purchase_price_input or 0),
                "purchase_price_includes_excise": bool(includes_excise),
                "purchase_price_ex_excise": float(purchase_ex or 0),
                "excise_per_unit": float(excise or 0),
                "packaging_cost_per_unit": float(packaging or 0),
                "other_direct_cost_per_unit": float(other_direct or 0),
                "updated_at": updated_at.isoformat() if hasattr(updated_at, "isoformat") and updated_at else "",
            }
        )
    return out


def upsert_lot_cost_record(raw: dict[str, Any]) -> dict[str, Any]:
    ensure_schema()
    lot_number = _text(raw.get("lot_number", raw.get("batchnummer", raw.get("lot", ""))))
    sku_code = _text(raw.get("sku_code", raw.get("sku", raw.get("artikelnummer", ""))))
    sku_id = _text(raw.get("sku_id", ""))
    supplier = _text(raw.get("supplier", raw.get("leverancier", "")))
    source_type = _text(raw.get("source_type", "")) or "purchase_invoice"
    source_ref = _text(raw.get("source_ref", raw.get("factuurnummer", "")))
    source_date = _parse_date(raw.get("source_date", raw.get("datum", raw.get("factuurdatum", ""))))
    includes_excise = _bool(raw.get("purchase_price_includes_excise", raw.get("prijs_inclusief_accijns", False)))
    purchase_input = _num(raw.get("purchase_price_input", raw.get("purchase_price", raw.get("inkoopprijs", 0))))
    excise_per_unit = _num(raw.get("excise_per_unit", raw.get("accijns_per_unit", 0)))
    purchase_ex = _num(raw.get("purchase_price_ex_excise", raw.get("inkoop_ex_accijns", 0)))
    if purchase_ex <= 0:
        purchase_ex = max(purchase_input - excise_per_unit, 0.0) if includes_excise else purchase_input
    record_id = _text(raw.get("id", "")) or _id_for(source_type, source_ref, supplier, lot_number, sku_id or sku_code)
    payload = dict(raw)
    normalized = {
        "id": record_id,
        "source_type": source_type,
        "source_ref": source_ref,
        "supplier": supplier,
        "lot_number": lot_number,
        "sku_id": sku_id,
        "sku_code": sku_code,
        "product_name": _text(raw.get("product_name", raw.get("product", ""))),
        "source_date": source_date.isoformat() if source_date else "",
        "quantity": _num(raw.get("quantity", raw.get("aantal", 0))),
        "purchase_price_input": purchase_input,
        "purchase_price_includes_excise": includes_excise,
        "purchase_price_ex_excise": purchase_ex,
        "excise_per_unit": excise_per_unit,
        "packaging_cost_per_unit": _num(raw.get("packaging_cost_per_unit", raw.get("verpakking_per_unit", 0))),
        "other_direct_cost_per_unit": _num(raw.get("other_direct_cost_per_unit", raw.get("extra_direct_per_unit", 0))),
    }
    if not lot_number:
        raise ValueError("LOT nummer ontbreekt.")
    if not sku_code and not sku_id:
        raise ValueError("SKU ontbreekt.")
    with postgres_storage.connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO lot_cost_records (
                    id, source_type, source_ref, supplier, lot_number, sku_id, sku_code, product_name, source_date,
                    quantity, purchase_price_input, purchase_price_includes_excise, purchase_price_ex_excise,
                    excise_per_unit, packaging_cost_per_unit, other_direct_cost_per_unit, payload, updated_at
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, %s)
                ON CONFLICT (id) DO UPDATE SET
                    source_type = EXCLUDED.source_type,
                    source_ref = EXCLUDED.source_ref,
                    supplier = EXCLUDED.supplier,
                    lot_number = EXCLUDED.lot_number,
                    sku_id = EXCLUDED.sku_id,
                    sku_code = EXCLUDED.sku_code,
                    product_name = EXCLUDED.product_name,
                    source_date = EXCLUDED.source_date,
                    quantity = EXCLUDED.quantity,
                    purchase_price_input = EXCLUDED.purchase_price_input,
                    purchase_price_includes_excise = EXCLUDED.purchase_price_includes_excise,
                    purchase_price_ex_excise = EXCLUDED.purchase_price_ex_excise,
                    excise_per_unit = EXCLUDED.excise_per_unit,
                    packaging_cost_per_unit = EXCLUDED.packaging_cost_per_unit,
                    other_direct_cost_per_unit = EXCLUDED.other_direct_cost_per_unit,
                    payload = EXCLUDED.payload,
                    updated_at = EXCLUDED.updated_at
                """,
                (
                    record_id,
                    source_type,
                    source_ref,
                    supplier,
                    lot_number,
                    sku_id,
                    sku_code,
                    normalized["product_name"],
                    source_date,
                    normalized["quantity"],
                    purchase_input,
                    includes_excise,
                    purchase_ex,
                    excise_per_unit,
                    normalized["packaging_cost_per_unit"],
                    normalized["other_direct_cost_per_unit"],
                    json.dumps(payload),
                    _now(),
                ),
            )
        if not postgres_storage.in_transaction():
            conn.commit()
    try:
        product_model_storage.upsert_purchase_lot_cost(normalized)
    except Exception:
        # Keep the legacy LOT-cost path available; the datamodel audit will surface sync gaps.
        pass
    return normalized


def find_lot_cost(*, lot_number: str, sku_code: str = "", sku_id: str = "") -> dict[str, Any] | None:
    ensure_schema()
    lot = _text(lot_number)
    if not lot:
        return None
    sku_code_text = _text(sku_code)
    sku_id_text = _text(sku_id)

    match_clauses: list[str] = []
    params: list[Any] = [lot]
    if sku_id_text:
        match_clauses.append("sku_id = %s")
        params.append(sku_id_text)
    if sku_code_text:
        match_clauses.append("sku_code = %s")
        params.append(sku_code_text)
    match_clauses.append("(COALESCE(sku_id, '') = '' AND COALESCE(sku_code, '') = '')")
    sku_where = " OR ".join(match_clauses)

    order_params: list[Any] = []
    if sku_id_text:
        order_params.append(sku_id_text)
    if sku_code_text:
        order_params.append(sku_code_text)

    order_parts: list[str] = []
    if sku_id_text:
        order_parts.append("(sku_id = %s) DESC")
    if sku_code_text:
        order_parts.append("(sku_code = %s) DESC")
    order_sql = ", ".join(order_parts + ["source_date DESC NULLS LAST", "updated_at DESC"])

    with postgres_storage.connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT id, supplier, purchase_price_ex_excise, excise_per_unit, packaging_cost_per_unit,
                       other_direct_cost_per_unit, source_type, source_ref
                FROM lot_cost_records
                WHERE LOWER(lot_number) = LOWER(%s)
                  AND ({sku_where})
                ORDER BY {order_sql}
                LIMIT 1
                """,
                tuple(params + order_params),
            )
            row = cur.fetchone()
    if not row:
        return None
    rid, supplier, purchase_ex, excise, packaging, other_direct, source_type, source_ref = row
    direct = float(purchase_ex or 0) + float(excise or 0) + float(packaging or 0) + float(other_direct or 0)
    return {
        "id": _text(rid),
        "supplier": _text(supplier),
        "purchase_price_ex_excise": float(purchase_ex or 0),
        "excise_per_unit": float(excise or 0),
        "packaging_cost_per_unit": float(packaging or 0),
        "other_direct_cost_per_unit": float(other_direct or 0),
        "cost_price_ex": direct,
        "source_type": _text(source_type),
        "source_ref": _text(source_ref),
    }


def find_lot_alias(*, douano_lot_number: str, sku_code: str = "", sku_id: str = "") -> dict[str, Any] | None:
    ensure_schema()
    douano_lot = _text(douano_lot_number)
    if not douano_lot:
        return None
    sku_code_text = _text(sku_code)
    sku_id_text = _text(sku_id)
    match_clauses: list[str] = []
    params: list[Any] = [douano_lot]
    if sku_id_text:
        match_clauses.append("sku_id = %s")
        params.append(sku_id_text)
    if sku_code_text:
        match_clauses.append("sku_code = %s")
        params.append(sku_code_text)
    match_clauses.append("(COALESCE(sku_id, '') = '' AND COALESCE(sku_code, '') = '')")
    sku_where = " OR ".join(match_clauses)

    order_params: list[Any] = []
    order_parts: list[str] = []
    if sku_id_text:
        order_parts.append("(sku_id = %s) DESC")
        order_params.append(sku_id_text)
    if sku_code_text:
        order_parts.append("(sku_code = %s) DESC")
        order_params.append(sku_code_text)
    order_sql = ", ".join(order_parts + ["updated_at DESC"])

    with postgres_storage.connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT id, sku_id, sku_code, douano_lot_number, internal_lot_number, reason, payload, updated_at
                FROM lot_alias_mappings
                WHERE LOWER(douano_lot_number) = LOWER(%s)
                  AND ({sku_where})
                ORDER BY {order_sql}
                LIMIT 1
                """,
                tuple(params + order_params),
            )
            row = cur.fetchone()
    if not row:
        return None
    rid, row_sku_id, row_sku_code, douano, internal, reason, payload, updated_at = row
    return {
        **(payload if isinstance(payload, dict) else {}),
        "id": _text(rid),
        "sku_id": _text(row_sku_id),
        "sku_code": _text(row_sku_code),
        "douano_lot_number": _text(douano),
        "internal_lot_number": _text(internal),
        "reason": _text(reason),
        "updated_at": updated_at.isoformat() if hasattr(updated_at, "isoformat") and updated_at else "",
    }


def upsert_lot_alias(raw: dict[str, Any]) -> dict[str, Any]:
    ensure_schema()
    sku_id = _text(raw.get("sku_id"))
    sku_code = _text(raw.get("sku_code", raw.get("sku")))
    douano_lot = _text(raw.get("douano_lot_number", raw.get("douano_lot", raw.get("lot_number"))))
    internal_lot = _text(raw.get("internal_lot_number", raw.get("internal_lot")))
    reason = _text(raw.get("reason"))
    if not douano_lot:
        raise ValueError("Douano LOT ontbreekt.")
    if not internal_lot:
        raise ValueError("Interne LOT ontbreekt.")
    if not sku_id and not sku_code:
        raise ValueError("SKU ontbreekt.")
    record_id = _text(raw.get("id")) or _id_for("lot_alias", sku_id, sku_code, douano_lot)
    payload = dict(raw)
    now = _now()
    with postgres_storage.connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id
                FROM lot_alias_mappings
                WHERE LOWER(COALESCE(sku_id, '')) = LOWER(%s)
                  AND LOWER(COALESCE(sku_code, '')) = LOWER(%s)
                  AND LOWER(douano_lot_number) = LOWER(%s)
                LIMIT 1
                """,
                (sku_id, sku_code, douano_lot),
            )
            existing = cur.fetchone()
            if existing:
                record_id = _text(existing[0])
                cur.execute(
                    """
                    UPDATE lot_alias_mappings
                    SET internal_lot_number = %s,
                        reason = %s,
                        payload = %s::jsonb,
                        updated_at = %s
                    WHERE id = %s
                    """,
                    (internal_lot, reason, json.dumps(payload), now, record_id),
                )
            else:
                cur.execute(
                    """
                    INSERT INTO lot_alias_mappings (
                        id, sku_id, sku_code, douano_lot_number, internal_lot_number, reason, payload, updated_at
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s::jsonb, %s)
                    """,
                    (record_id, sku_id, sku_code, douano_lot, internal_lot, reason, json.dumps(payload), now),
                )
        if not postgres_storage.in_transaction():
            conn.commit()
    return {
        "id": record_id,
        "sku_id": sku_id,
        "sku_code": sku_code,
        "douano_lot_number": douano_lot,
        "internal_lot_number": internal_lot,
        "reason": reason,
        "updated_at": now.isoformat(),
    }


def delete_lot_alias(alias_id: str) -> bool:
    ensure_schema()
    rid = _text(alias_id)
    if not rid:
        return False
    with postgres_storage.connect() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM lot_alias_mappings WHERE id = %s", (rid,))
            deleted = int(cur.rowcount or 0)
        if not postgres_storage.in_transaction():
            conn.commit()
    return deleted > 0


_LOT_KEYS = {"lotnummer", "lotnumber", "lot", "batchnummer", "batchnumber", "batch"}


def _lot_key_name(value: Any) -> str:
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


def _version_lot_numbers(version: dict[str, Any]) -> list[str]:
    lots: list[str] = []
    seen: set[str] = set()

    def collect(value: Any, parent_key: str = "") -> None:
        if isinstance(value, dict):
            for key, child in value.items():
                key_text = _lot_key_name(key)
                if key_text in _LOT_KEYS:
                    text = _text(child)
                    if text and text.lower() not in seen:
                        seen.add(text.lower())
                        lots.append(text)
                    continue
                collect(child, key_text)
            return
        if isinstance(value, list):
            for child in value:
                collect(child, parent_key)

    collect(version)
    return lots


def _version_lot_number(version: dict[str, Any]) -> str:
    lots = _version_lot_numbers(version)
    return lots[0] if lots else ""


def _lot_exact_key(value: Any) -> str:
    """Canonical exact LOT comparison key.

    Douano LOT values are the source of truth for sales rows. This key removes
    harmless formatting characters, but intentionally does not treat letter O
    and digit 0 as the same value.
    """
    return "".join(ch for ch in _text(value).upper() if ch.isalnum())


def _lot_near_key(value: Any) -> str:
    """Diagnostic key for likely typing mistakes such as PO/P0.

    Near matches are only used for warnings and correction suggestions. They
    must not silently become the canonical LOT match.
    """
    return _lot_exact_key(value).replace("O", "0")


def _version_lot_candidates_by_sku(*, year: int = 0) -> dict[str, list[dict[str, Any]]]:
    try:
        from app.domain import cost_versions_storage

        return cost_versions_storage.load_lot_candidates_by_sku(year=int(year or 0), limit=50000)
    except Exception:
        return {}


def _replace_lot_values(value: Any, *, from_lot: str, to_lot: str) -> tuple[Any, int]:
    from_key = _lot_exact_key(from_lot)
    changed = 0
    if isinstance(value, dict):
        out: dict[str, Any] = {}
        for key, child in value.items():
            if _lot_key_name(key) in _LOT_KEYS and _lot_exact_key(child) == from_key:
                out[key] = to_lot
                changed += 1
                continue
            new_child, child_changed = _replace_lot_values(child, from_lot=from_lot, to_lot=to_lot)
            out[key] = new_child
            changed += child_changed
        return out, changed
    if isinstance(value, list):
        out_list: list[Any] = []
        for child in value:
            new_child, child_changed = _replace_lot_values(child, from_lot=from_lot, to_lot=to_lot)
            out_list.append(new_child)
            changed += child_changed
        return out_list, changed
    return value, 0


def correct_internal_lot_to_douano(raw: dict[str, Any]) -> dict[str, Any]:
    ensure_schema()
    sku_id = _text(raw.get("sku_id"))
    sku_code = _text(raw.get("sku_code", raw.get("sku")))
    douano_lot = _text(raw.get("douano_lot_number", raw.get("douano_lot")))
    internal_lot = _text(raw.get("internal_lot_number", raw.get("internal_lot")))
    raw_version_ids = raw.get("internal_version_ids", raw.get("version_ids", []))
    version_ids = [
        _text(value)
        for value in (raw_version_ids if isinstance(raw_version_ids, list) else [raw_version_ids])
        if _text(value)
    ]
    if not douano_lot:
        raise ValueError("Douano LOT ontbreekt.")
    if not internal_lot:
        raise ValueError("Interne LOT ontbreekt.")
    if _lot_exact_key(douano_lot) == _lot_exact_key(internal_lot):
        return {"updated_cost_versions": 0, "updated_lot_cost_records": 0, "deleted_aliases": 0, "message": "LOTs zijn al gelijk."}

    from app.domain import cost_versions_storage

    cost_versions_storage.ensure_schema()
    updated_versions = 0
    updated_lot_rows = 0
    updated_lot_records = 0
    deleted_aliases = 0
    now = _now()
    with postgres_storage.connect() as conn:
        with conn.cursor() as cur:
            if version_ids:
                placeholders = ", ".join(["%s"] * len(version_ids))
                cur.execute(
                    f"""
                    SELECT DISTINCT v.id, v.payload
                    FROM cost_versions v
                    WHERE v.id IN ({placeholders})
                    """,
                    tuple(version_ids),
                )
            else:
                cur.execute(
                    """
                    SELECT DISTINCT v.id, v.payload
                    FROM cost_versions v
                    JOIN cost_version_lots l ON l.version_id = v.id
                    LEFT JOIN cost_version_sku_rows r ON r.version_id = v.id
                    WHERE LOWER(l.lot_number) = LOWER(%s)
                      AND (
                        (%s <> '' AND r.sku_id = %s)
                        OR (%s <> '' AND EXISTS (
                            SELECT 1
                            FROM lot_cost_records c
                            WHERE LOWER(c.sku_code) = LOWER(%s)
                              AND LOWER(c.lot_number) = LOWER(%s)
                        ))
                        OR (%s = '' AND %s = '')
                      )
                    """,
                    (internal_lot, sku_id, sku_id, sku_code, sku_code, internal_lot, sku_id, sku_code),
                )
            selected_version_ids: list[str] = []
            for version_id, payload in cur.fetchall() or []:
                version_text = _text(version_id)
                if version_text:
                    selected_version_ids.append(version_text)
                version_payload = payload if isinstance(payload, dict) else {}
                new_payload, changed = _replace_lot_values(version_payload, from_lot=internal_lot, to_lot=douano_lot)
                if changed <= 0:
                    continue
                cur.execute(
                    """
                    UPDATE cost_versions
                    SET payload = %s::jsonb,
                        updated_at = %s,
                        updated_at_ts = %s
                    WHERE id = %s
                    """,
                    (json.dumps(new_payload), now, now, _text(version_id)),
                )
                updated_versions += 1

            if selected_version_ids:
                placeholders = ", ".join(["%s"] * len(selected_version_ids))
                # If the corrected Douano LOT already exists on the same version, remove that duplicate first.
                cur.execute(
                    f"""
                    DELETE FROM cost_version_lots target
                    WHERE target.version_id IN ({placeholders})
                      AND LOWER(target.lot_number) = LOWER(%s)
                      AND EXISTS (
                        SELECT 1
                        FROM cost_version_lots source
                        WHERE source.version_id = target.version_id
                          AND LOWER(source.lot_number) = LOWER(%s)
                          AND source.id <> target.id
                      )
                    """,
                    (*selected_version_ids, douano_lot, internal_lot),
                )
                cur.execute(
                    f"""
                    UPDATE cost_version_lots
                    SET lot_number = %s,
                        updated_at_ts = %s
                    WHERE version_id IN ({placeholders})
                      AND LOWER(lot_number) = LOWER(%s)
                    """,
                    (douano_lot, now, *selected_version_ids, internal_lot),
                )
                updated_lot_rows = int(cur.rowcount or 0)

            cur.execute(
                """
                UPDATE lot_cost_records
                SET lot_number = %s,
                    payload = jsonb_set(payload, '{lot_number}', to_jsonb(%s::text), true),
                    updated_at = %s
                WHERE LOWER(lot_number) = LOWER(%s)
                  AND (
                    COALESCE(sku_id, '') = ''
                    OR (%s <> '' AND sku_id = %s)
                    OR (%s <> '' AND LOWER(sku_code) = LOWER(%s))
                  )
                """,
                (douano_lot, douano_lot, now, internal_lot, sku_id, sku_id, sku_code, sku_code),
            )
            updated_lot_records = int(cur.rowcount or 0)

            cur.execute(
                """
                DELETE FROM lot_alias_mappings
                WHERE LOWER(douano_lot_number) = LOWER(%s)
                  AND LOWER(internal_lot_number) = LOWER(%s)
                  AND (
                    COALESCE(sku_id, '') = ''
                    OR (%s <> '' AND sku_id = %s)
                    OR (%s <> '' AND LOWER(sku_code) = LOWER(%s))
                  )
                """,
                (douano_lot, internal_lot, sku_id, sku_id, sku_code, sku_code),
            )
            deleted_aliases = int(cur.rowcount or 0)
        if not postgres_storage.in_transaction():
            conn.commit()

    return {
        "updated_cost_versions": updated_versions,
        "updated_cost_version_lots": updated_lot_rows,
        "updated_lot_cost_records": updated_lot_records,
        "deleted_aliases": deleted_aliases,
        "douano_lot_number": douano_lot,
        "previous_internal_lot_number": internal_lot,
    }


def list_lot_master_reconciliation(*, year: int = 0, limit: int = 5000) -> list[dict[str, Any]]:
    ensure_schema()
    lim = max(1, min(int(limit or 5000), 20000))
    version_candidates = _version_lot_candidates_by_sku(year=int(year or 0))
    where = "WHERE COALESCE(NULLIF(a.lot_number, ''), '') <> ''"
    params: list[Any] = []
    if int(year or 0) > 0:
        where += " AND a.movement_date >= %s::date AND a.movement_date < %s::date"
        params.extend([f"{int(year)}-01-01", f"{int(year) + 1}-01-01"])

    sku_meta: dict[str, dict[str, Any]] = {}
    douano_by_sku: dict[str, list[dict[str, Any]]] = {}
    beer_names: dict[str, str] = {}
    try:
        from app.domain import dataset_store

        beers = dataset_store.load_dataset("bieren")
        if isinstance(beers, list):
            for beer in beers:
                if isinstance(beer, dict):
                    bid = _text(beer.get("id"))
                    name = _text(beer.get("naam", beer.get("name", beer.get("biernaam", ""))))
                    if bid and name:
                        beer_names[bid] = name
    except Exception:
        beer_names = {}
    with postgres_storage.connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT s.id, COALESCE(s.code, '') AS code, COALESCE(s.name, '') AS name,
                       COALESCE(s.beer_id, '') AS beer_id,
                       COALESCE(a.name, '') AS article_name
                FROM skus s
                LEFT JOIN articles a ON a.id = COALESCE(NULLIF(s.article_id, ''), NULLIF(s.format_article_id, ''))
                LIMIT %s
                """,
                (lim,),
            )
            for sid, code, name, beer_id, article_name in cur.fetchall() or []:
                sku_id = _text(sid)
                if not sku_id:
                    continue
                beer_id_text = _text(beer_id)
                sku_meta[sku_id] = {
                    "sku_id": sku_id,
                    "sku_code": _text(code),
                    "sku_name": _text(name) or _text(article_name) or sku_id,
                    "style_id": beer_id_text or sku_id,
                    "style_name": beer_names.get(beer_id_text, "") or _text(name) or _text(article_name) or sku_id,
                }

            sku_id_by_code = {
                _text(meta.get("sku_code")).lower(): sku_id
                for sku_id, meta in sku_meta.items()
                if _text(meta.get("sku_code"))
            }

            cur.execute(
                f"""
                SELECT
                    COALESCE(m.sku_id, '') AS sku_id,
                    a.sku_code,
                    a.lot_number,
                    MAX(a.product_name) AS product_name,
                    COUNT(*)::int AS rows,
                    MAX(a.movement_date) AS last_movement_date
                FROM sales_lot_allocations a
                LEFT JOIN douano_products p ON LOWER(p.sku) = LOWER(a.sku_code)
                LEFT JOIN douano_product_mapping m ON m.douano_product_id = p.product_id
                {where}
                GROUP BY COALESCE(m.sku_id, ''), a.sku_code, a.lot_number
                ORDER BY rows DESC
                LIMIT %s
                """,
                tuple(params + [lim]),
            )
            for sku_id, sku_code, lot, product_name, rows, last_movement_date in cur.fetchall() or []:
                sku_id_text = _text(sku_id)
                sku_code_text = _text(sku_code)
                if not sku_id_text and sku_code_text:
                    sku_id_text = sku_id_by_code.get(sku_code_text.lower(), "")
                if sku_id_text and sku_id_text not in sku_meta:
                    sku_meta[sku_id_text] = {
                        "sku_id": sku_id_text,
                        "sku_code": sku_code_text,
                        "sku_name": _text(product_name) or sku_id_text,
                        "style_id": sku_id_text,
                        "style_name": _text(product_name) or sku_id_text,
                    }
                key = sku_id_text or f"sku-code:{sku_code_text}"
                douano_by_sku.setdefault(key, []).append(
                    {
                        "lot_number": _text(lot),
                        "sku_code": sku_code_text,
                        "product_name": _text(product_name),
                        "rows": int(rows or 0),
                        "last_movement_date": last_movement_date.isoformat() if last_movement_date else "",
                    }
                )

            cur.execute(
                """
                SELECT DISTINCT sku_id, sku_code, lot_number, source_type, source_ref, source_date
                FROM lot_cost_records
                WHERE COALESCE(NULLIF(lot_number, ''), '') <> ''
                ORDER BY source_date DESC NULLS LAST, lot_number
                LIMIT %s
                """,
                (lim,),
            )
            for sku_id, sku_code, lot, source_type, source_ref, source_date in cur.fetchall() or []:
                sku_id_text = _text(sku_id)
                sku_code_text = _text(sku_code)
                key = sku_id_text or f"sku-code:{sku_code_text}"
                item = {
                    "lot_number": _text(lot),
                    "source": _text(source_type) or "lot_cost",
                    "label": _text(source_ref) or _text(source_type) or "LOT kostprijs",
                    "source_date": source_date.isoformat() if source_date else "",
                    "version_ids": [],
                    "labels": [],
                }
                if sku_id_text:
                    version_candidates.setdefault(sku_id_text, []).append(item)
                elif sku_code_text:
                    version_candidates.setdefault(key, []).append(item)

    groups: dict[str, dict[str, Any]] = {}
    all_keys = sorted(set(version_candidates.keys()) | set(douano_by_sku.keys()))
    for key in all_keys:
        meta = sku_meta.get(key, {})
        if not meta and key.startswith("sku-code:"):
            meta = {
                "sku_id": "",
                "sku_code": key.replace("sku-code:", "", 1),
                "sku_name": key.replace("sku-code:", "", 1),
                "style_id": key,
                "style_name": key.replace("sku-code:", "", 1),
            }
        style_id = _text(meta.get("style_id")) or key
        group = groups.setdefault(
            style_id,
            {
                "style_id": style_id,
                "style_name": _text(meta.get("style_name")) or style_id,
                "rows": [],
                "summary": {"internal_lots": 0, "douano_lots": 0, "matched": 0, "near_match": 0, "missing": 0, "douano_only": 0},
            },
        )
        douano_lots = douano_by_sku.get(key, [])
        internal_lots = version_candidates.get(key, [])
        used_douano: set[str] = set()
        for internal in internal_lots:
            internal_lot = _text(internal.get("lot_number"))
            if not internal_lot:
                continue
            exact = next((row for row in douano_lots if _lot_exact_key(row.get("lot_number")) == _lot_exact_key(internal_lot)), None)
            near = next(
                (
                    row
                    for row in douano_lots
                    if _lot_exact_key(row.get("lot_number")) != _lot_exact_key(internal_lot)
                    and _lot_near_key(row.get("lot_number")) == _lot_near_key(internal_lot)
                ),
                None,
            )
            match = exact or near
            if match:
                used_douano.add(_lot_exact_key(match.get("lot_number")))
            status = "matched" if exact else "near_match" if near else "missing_douano"
            group["rows"].append(
                {
                    "sku_id": _text(meta.get("sku_id")) if meta else (key if not key.startswith("sku-code:") else ""),
                    "sku_code": _text(meta.get("sku_code")) or _text((match or {}).get("sku_code")),
                    "sku_name": _text(meta.get("sku_name")) or _text((match or {}).get("product_name")) or key,
                    "internal_lot_number": internal_lot,
                    "internal_label": _text(internal.get("label")),
                    "internal_labels": internal.get("labels") if isinstance(internal.get("labels"), list) else [],
                    "internal_source": _text(internal.get("source")),
                    "internal_version_ids": internal.get("version_ids") if isinstance(internal.get("version_ids"), list) else [],
                    "douano_lot_number": _text((match or {}).get("lot_number")),
                    "douano_options": douano_lots,
                    "status": status,
                    "rows": int((match or {}).get("rows", 0) or 0),
                    "last_movement_date": _text((match or {}).get("last_movement_date")),
                }
            )
            group["summary"]["internal_lots"] += 1
            group["summary"]["matched" if exact else "near_match" if near else "missing"] += 1
        for douano in douano_lots:
            if _lot_exact_key(douano.get("lot_number")) in used_douano:
                continue
            has_near_internal = any(
                _lot_near_key(row.get("lot_number")) == _lot_near_key(douano.get("lot_number"))
                for row in internal_lots
            )
            if has_near_internal:
                continue
            group["rows"].append(
                {
                    "sku_id": _text(meta.get("sku_id")) if meta else (key if not key.startswith("sku-code:") else ""),
                    "sku_code": _text(meta.get("sku_code")) or _text(douano.get("sku_code")),
                    "sku_name": _text(meta.get("sku_name")) or _text(douano.get("product_name")) or key,
                    "internal_lot_number": "",
                    "internal_label": "",
                    "internal_labels": [],
                    "internal_source": "",
                    "internal_version_ids": [],
                    "douano_lot_number": _text(douano.get("lot_number")),
                    "douano_options": douano_lots,
                    "status": "douano_only",
                    "rows": int(douano.get("rows", 0) or 0),
                    "last_movement_date": _text(douano.get("last_movement_date")),
                }
            )
            group["summary"]["douano_only"] += 1
        group["summary"]["douano_lots"] += len(douano_lots)

    out = list(groups.values())
    for group in out:
        group["rows"].sort(key=lambda row: (str(row.get("sku_name", "")), str(row.get("internal_lot_number", "")), str(row.get("douano_lot_number", ""))))
    out.sort(key=lambda row: (row["summary"].get("near_match", 0) + row["summary"].get("missing", 0) + row["summary"].get("douano_only", 0) == 0, str(row.get("style_name", ""))))
    return out


def list_lot_reconciliation(*, year: int = 0, limit: int = 500) -> list[dict[str, Any]]:
    ensure_schema()
    lim = max(1, min(int(limit or 500), 5000))
    where = "WHERE COALESCE(NULLIF(a.lot_number, ''), '') <> ''"
    params: list[Any] = []
    if int(year or 0) > 0:
        where += " AND a.movement_date >= %s::date AND a.movement_date < %s::date"
        params.extend([f"{int(year)}-01-01", f"{int(year) + 1}-01-01"])
    version_candidates = _version_lot_candidates_by_sku(year=int(year or 0))
    with postgres_storage.connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                WITH sales AS (
                    SELECT
                        a.sku_code,
                        a.lot_number AS douano_lot_number,
                        MAX(a.product_name) AS douano_product_name,
                        COUNT(*)::int AS rows,
                        MAX(a.movement_date) AS last_movement_date,
                        COALESCE(MAX(m.sku_id), '') AS sku_id
                    FROM sales_lot_allocations a
                    LEFT JOIN douano_products p ON LOWER(p.sku) = LOWER(a.sku_code)
                    LEFT JOIN douano_product_mapping m ON m.douano_product_id = p.product_id
                    {where}
                    GROUP BY a.sku_code, a.lot_number
                )
                SELECT
                    s.sku_code,
                    s.douano_lot_number,
                    s.douano_product_name,
                    s.rows,
                    s.last_movement_date,
                    s.sku_id,
                    COALESCE(sku.name, art.name, s.douano_product_name, '') AS sku_name,
                    COALESCE(sku.beer_id, '') AS group_key,
                    alias.id AS alias_id,
                    alias.internal_lot_number,
                    alias.reason,
                    COUNT(DISTINCT direct.id)::int AS direct_cost_records,
                    COUNT(DISTINCT mapped.id)::int AS mapped_cost_records
                FROM sales s
                LEFT JOIN skus sku ON sku.id = s.sku_id
                LEFT JOIN articles art ON art.id = COALESCE(NULLIF(sku.article_id, ''), NULLIF(sku.format_article_id, ''))
                LEFT JOIN lot_alias_mappings alias
                  ON LOWER(alias.douano_lot_number) = LOWER(s.douano_lot_number)
                 AND (
                    (COALESCE(alias.sku_id, '') <> '' AND alias.sku_id = s.sku_id)
                    OR (COALESCE(alias.sku_code, '') <> '' AND LOWER(alias.sku_code) = LOWER(s.sku_code))
                    OR (COALESCE(alias.sku_id, '') = '' AND COALESCE(alias.sku_code, '') = '')
                 )
                LEFT JOIN lot_cost_records direct
                  ON LOWER(direct.lot_number) = LOWER(s.douano_lot_number)
                 AND (
                    COALESCE(direct.sku_id, '') = ''
                    OR direct.sku_id = s.sku_id
                    OR LOWER(direct.sku_code) = LOWER(s.sku_code)
                 )
                LEFT JOIN lot_cost_records mapped
                  ON LOWER(mapped.lot_number) = LOWER(alias.internal_lot_number)
                 AND (
                    COALESCE(mapped.sku_id, '') = ''
                    OR mapped.sku_id = s.sku_id
                    OR LOWER(mapped.sku_code) = LOWER(s.sku_code)
                 )
                GROUP BY
                    s.sku_code, s.douano_lot_number, s.douano_product_name, s.rows, s.last_movement_date,
                    s.sku_id, sku.name, art.name, sku.beer_id, alias.id, alias.internal_lot_number, alias.reason
                ORDER BY
                    CASE
                        WHEN COUNT(DISTINCT direct.id) > 0 THEN 3
                        WHEN alias.id IS NOT NULL AND (COUNT(DISTINCT mapped.id) > 0 OR COALESCE(alias.internal_lot_number, '') <> '') THEN 2
                        ELSE 1
                    END,
                    s.rows DESC,
                    s.sku_code,
                    s.douano_lot_number
                LIMIT %s
                """,
                tuple(params + [lim]),
            )
            rows = cur.fetchall() or []
            sku_keys = sorted({_text(row[5]) for row in rows if _text(row[5])})
            sku_codes = sorted({_text(row[0]) for row in rows if _text(row[0])})
            cost_candidates: dict[tuple[str, str], list[dict[str, Any]]] = {}
            if sku_keys or sku_codes:
                cur.execute(
                    """
                    SELECT DISTINCT sku_id, sku_code, lot_number, source_type, source_ref, source_date
                    FROM lot_cost_records
                    WHERE COALESCE(NULLIF(lot_number, ''), '') <> ''
                      AND (
                        (%s::text[] <> '{}'::text[] AND sku_id = ANY(%s::text[]))
                        OR (%s::text[] <> '{}'::text[] AND sku_code = ANY(%s::text[]))
                      )
                    ORDER BY source_date DESC NULLS LAST, lot_number
                    """,
                    (sku_keys, sku_keys, sku_codes, sku_codes),
                )
                for sku_id, sku_code, lot, source_type, source_ref, source_date in cur.fetchall() or []:
                    item = {
                        "lot_number": _text(lot),
                        "source": _text(source_type) or "lot_cost",
                        "label": _text(source_ref) or _text(source_type) or "LOT kostprijs",
                        "source_date": source_date.isoformat() if source_date else "",
                    }
                    for key in ((_text(sku_id), ""), ("", _text(sku_code))):
                        if key[0] or key[1]:
                            cost_candidates.setdefault(key, []).append(item)

    out: list[dict[str, Any]] = []
    for row in rows:
        (
            sku_code,
            douano_lot,
            douano_product_name,
            count_rows,
            last_movement_date,
            sku_id,
            sku_name,
            group_key,
            alias_id,
            internal_lot,
            reason,
            direct_count,
            mapped_count,
        ) = row
        sku_id_text = _text(sku_id)
        sku_code_text = _text(sku_code)
        candidates: list[dict[str, Any]] = []
        seen_lots: set[str] = set()
        douano_lot_text = _text(douano_lot)
        for candidate in [
            *cost_candidates.get((sku_id_text, ""), []),
            *cost_candidates.get(("", sku_code_text), []),
            *version_candidates.get(sku_id_text, []),
        ]:
            lot = _text(candidate.get("lot_number"))
            if not lot or lot.lower() in seen_lots:
                continue
            seen_lots.add(lot.lower())
            exact = _lot_exact_key(lot) == _lot_exact_key(douano_lot_text)
            near = not exact and _lot_near_key(lot) == _lot_near_key(douano_lot_text)
            candidates.append({**candidate, "exact_match": exact, "near_match": near})
        exact_candidate = next((candidate for candidate in candidates if bool(candidate.get("exact_match"))), None)
        near_candidate = next((candidate for candidate in candidates if bool(candidate.get("near_match"))), None)
        candidate_lots = {str(candidate.get("lot_number", "") or "").strip().lower() for candidate in candidates}
        if int(direct_count or 0) > 0 or exact_candidate is not None:
            status = "direct"
        elif _text(alias_id):
            status = "mapped" if int(mapped_count or 0) > 0 or _text(internal_lot).lower() in candidate_lots else "mapped_missing_cost"
        elif near_candidate is not None:
            status = "near_match"
            internal_lot = _text(near_candidate.get("lot_number"))
        else:
            status = "missing"
        out.append(
            {
                "sku_code": sku_code_text,
                "sku_id": sku_id_text,
                "sku_name": _text(sku_name),
                "group_key": _text(group_key),
                "douano_product_name": _text(douano_product_name),
                "douano_lot_number": _text(douano_lot),
                "internal_lot_number": _text(internal_lot),
                "suggested_internal_lot_number": _text(near_candidate.get("lot_number")) if near_candidate else "",
                "alias_id": _text(alias_id),
                "reason": _text(reason),
                "rows": int(count_rows or 0),
                "last_movement_date": last_movement_date.isoformat() if last_movement_date else "",
                "direct_cost_records": int(direct_count or 0),
                "mapped_cost_records": int(mapped_count or 0),
                "status": status,
                "internal_lot_candidates": candidates[:25],
            }
        )
    return out


def find_sales_lot(*, transaction_number: str, sku_code: str) -> dict[str, Any] | None:
    ensure_schema()
    tx = _text(transaction_number)
    sku = _text(sku_code)
    if not tx or not sku:
        return None
    with postgres_storage.connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT lot_number, quantity, stock_value_per_unit, excise_per_unit
                FROM sales_lot_allocations
                WHERE transaction_number = %s AND sku_code = %s
                ORDER BY ABS(quantity) DESC, updated_at DESC
                LIMIT 1
                """,
                (tx, sku),
            )
            row = cur.fetchone()
    if not row:
        return None
    lot_number, quantity, stock_value, excise = row
    return {
        "lot_number": _text(lot_number),
        "quantity": float(quantity or 0),
        "stock_value_per_unit": float(stock_value or 0),
        "excise_per_unit": float(excise or 0),
    }


def find_sales_lot_any(*, transaction_numbers: list[str], sku_code: str) -> dict[str, Any] | None:
    ensure_schema()
    txs = [_text(tx) for tx in transaction_numbers if _text(tx)]
    sku = _text(sku_code)
    if not txs or not sku:
        return None
    with postgres_storage.connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT lot_number, quantity, stock_value_per_unit, excise_per_unit, transaction_number
                FROM sales_lot_allocations
                WHERE transaction_number = ANY(%s) AND sku_code = %s
                ORDER BY CASE WHEN COALESCE(NULLIF(lot_number, ''), '') = '' THEN 1 ELSE 0 END,
                         ABS(quantity) DESC,
                         updated_at DESC
                LIMIT 1
                """,
                (txs, sku),
            )
            row = cur.fetchone()
    if not row:
        return None
    lot_number, quantity, stock_value, excise, transaction_number = row
    return {
        "lot_number": _text(lot_number),
        "quantity": float(quantity or 0),
        "stock_value_per_unit": float(stock_value or 0),
        "excise_per_unit": float(excise or 0),
        "transaction_number": _text(transaction_number),
    }


def _column_key(value: Any) -> str:
    text = _text(value).lower()
    replacements = {
        "datum": "date",
        "date": "date",
        "source date": "source_date",
        "factuurdatum": "source_date",
        "artikelnummer": "sku_code",
        "sku code": "sku_code",
        "sku": "sku_code",
        "product": "product_name",
        "productnaam": "product_name",
        "batchnummer": "lot_number",
        "lot": "lot_number",
        "lotnummer": "lot_number",
        "lot nummer": "lot_number",
        "transactienummer": "transaction_number",
        "transactionnumber": "transaction_number",
        "bedrijf": "company_name",
        "aantal": "quantity",
        "aantal ref": "quantity",
        "aantal referentie": "quantity",
        "supplier": "supplier",
        "leverancier": "supplier",
        "inkoopprijs": "purchase_price_input",
        "purchase price": "purchase_price_input",
        "accijns": "excise_per_unit",
        "accijns per unit": "excise_per_unit",
        "inclusief accijns": "purchase_price_includes_excise",
        "prijs inclusief accijns": "purchase_price_includes_excise",
        "waarde per eenheid": "stock_value_per_unit",
        "accijns per eenheid": "excise_per_unit",
        "transactietype": "movement_type",
        "reden": "movement_reason",
        "voorraadbeweging": "stock_movement",
    }
    normalized = " ".join(text.replace("_", " ").replace("-", " ").split())
    return replacements.get(normalized, normalized.replace(" ", "_"))


def _validate_stock_history_headers(keys: list[str]) -> None:
    present = {key for key in keys if key}
    missing = [label for label in STOCK_HISTORY_COLUMNS if _column_key(label) not in present]
    if missing:
        raise ValueError("Voorraadhistoriek mist kolommen: " + ", ".join(missing))


def _xlsx_column_index(ref: str) -> int:
    letters = "".join(ch for ch in ref if ch.isalpha()).upper()
    idx = 0
    for ch in letters:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return max(idx - 1, 0)


def _read_xlsx_basic(content: bytes) -> list[dict[str, Any]]:
    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            shared: list[str] = []
            if "xl/sharedStrings.xml" in archive.namelist():
                root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
                for item in root.findall("x:si", ns):
                    shared.append("".join(node.text or "" for node in item.findall(".//x:t", ns)))
            root = ET.fromstring(archive.read("xl/worksheets/sheet1.xml"))
    except (KeyError, zipfile.BadZipFile, ET.ParseError) as exc:
        raise ValueError("Excel-bestand kon niet worden gelezen. Download het voorbeeld opnieuw of sla het bestand opnieuw op als .xlsx.") from exc

    rows: list[list[Any]] = []
    for row_el in root.findall(".//x:sheetData/x:row", ns):
        values: list[Any] = []
        for cell in row_el.findall("x:c", ns):
            ref = str(cell.attrib.get("r", ""))
            col_idx = _xlsx_column_index(ref)
            while len(values) <= col_idx:
                values.append(None)
            cell_type = cell.attrib.get("t")
            value = ""
            if cell_type == "inlineStr":
                value = "".join(node.text or "" for node in cell.findall(".//x:t", ns))
            else:
                raw = cell.find("x:v", ns)
                value = raw.text if raw is not None and raw.text is not None else ""
                if cell_type == "s":
                    try:
                        value = shared[int(value)]
                    except (ValueError, IndexError):
                        value = ""
                elif cell_type not in {"str", "b"}:
                    try:
                        value = float(value) if "." in value else int(value)
                    except (TypeError, ValueError):
                        pass
            values[col_idx] = value
        if any(_text(v) for v in values):
            rows.append(values)

    if not rows:
        return []
    headers = rows[0]
    keys = [_column_key(cell) for cell in headers]
    _validate_stock_history_headers(keys)
    out: list[dict[str, Any]] = []
    for values in rows[1:]:
        row = {
            keys[idx]: values[idx] if idx < len(values) else None
            for idx in range(len(keys))
            if keys[idx] in STOCK_HISTORY_REQUIRED_KEYS
        }
        if any(_text(v) for v in row.values()):
            out.append(row)
    return out


def _read_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    sample = text[:2048]
    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    _validate_stock_history_headers([_column_key(k) for k in (reader.fieldnames or [])])
    return [
        {key: value for key, value in ((_column_key(k), v) for k, v in row.items()) if key in STOCK_HISTORY_REQUIRED_KEYS}
        for row in reader
    ]


def _read_xlsx(content: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        return _read_xlsx_basic(content)
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except zipfile.BadZipFile as exc:
        raise ValueError("Excel-bestand kon niet worden gelezen. Download het voorbeeld opnieuw of sla het bestand opnieuw op als .xlsx.") from exc
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return []
    keys = [_column_key(cell) for cell in headers]
    _validate_stock_history_headers(keys)
    out: list[dict[str, Any]] = []
    for values in rows_iter:
        row = {
            keys[idx]: values[idx] if idx < len(values) else None
            for idx in range(len(keys))
            if keys[idx] in STOCK_HISTORY_REQUIRED_KEYS
        }
        if any(_text(v) for v in row.values()):
            out.append(row)
    return out


def _load_stock_history_rows(content: bytes, filename: str) -> list[dict[str, Any]]:
    lower = _text(filename).lower()
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return _read_xlsx(content)
    if lower.endswith(".csv") or lower.endswith(".txt"):
        return _read_csv(content)
    raise ValueError("Ondersteund formaat: .xlsx, .xlsm, .csv of .txt.")


def _validate_opening_lot_headers(keys: list[str]) -> None:
    present = {key for key in keys if key}
    missing = [label for label in OPENING_LOT_COLUMNS if _column_key(label) not in present]
    if missing:
        raise ValueError("Opening LOT bestand mist kolommen: " + ", ".join(missing))


def _read_opening_lot_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    sample = text[:2048]
    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    _validate_opening_lot_headers([_column_key(k) for k in (reader.fieldnames or [])])
    return [
        {key: value for key, value in ((_column_key(k), v) for k, v in row.items()) if key in OPENING_LOT_REQUIRED_KEYS}
        for row in reader
    ]


def _read_opening_lot_xlsx_basic(content: bytes) -> list[dict[str, Any]]:
    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            shared: list[str] = []
            if "xl/sharedStrings.xml" in archive.namelist():
                root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
                for item in root.findall("x:si", ns):
                    shared.append("".join(node.text or "" for node in item.findall(".//x:t", ns)))
            root = ET.fromstring(archive.read("xl/worksheets/sheet1.xml"))
    except (KeyError, zipfile.BadZipFile, ET.ParseError) as exc:
        raise ValueError("Opening LOT Excel-bestand kon niet worden gelezen. Download het voorbeeld opnieuw of sla het bestand opnieuw op als .xlsx.") from exc

    rows: list[list[Any]] = []
    for row_el in root.findall(".//x:sheetData/x:row", ns):
        values: list[Any] = []
        for cell in row_el.findall("x:c", ns):
            ref = str(cell.attrib.get("r", ""))
            col_idx = _xlsx_column_index(ref)
            while len(values) <= col_idx:
                values.append(None)
            cell_type = cell.attrib.get("t")
            value = ""
            if cell_type == "inlineStr":
                value = "".join(node.text or "" for node in cell.findall(".//x:t", ns))
            else:
                raw = cell.find("x:v", ns)
                value = raw.text if raw is not None and raw.text is not None else ""
                if cell_type == "s":
                    try:
                        value = shared[int(value)]
                    except (ValueError, IndexError):
                        value = ""
                elif cell_type not in {"str", "b"}:
                    try:
                        value = float(value) if "." in value else int(value)
                    except (TypeError, ValueError):
                        pass
            values[col_idx] = value
        if any(_text(v) for v in values):
            rows.append(values)

    if not rows:
        return []
    headers = rows[0]
    keys = [_column_key(cell) for cell in headers]
    _validate_opening_lot_headers(keys)
    out: list[dict[str, Any]] = []
    for values in rows[1:]:
        row = {
            keys[idx]: values[idx] if idx < len(values) else None
            for idx in range(len(keys))
            if keys[idx] in OPENING_LOT_REQUIRED_KEYS
        }
        if any(_text(v) for v in row.values()):
            out.append(row)
    return out


def _read_opening_lot_xlsx(content: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        return _read_opening_lot_xlsx_basic(content)
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except zipfile.BadZipFile as exc:
        raise ValueError("Opening LOT Excel-bestand kon niet worden gelezen. Download het voorbeeld opnieuw of sla het bestand opnieuw op als .xlsx.") from exc
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return []
    keys = [_column_key(cell) for cell in headers]
    _validate_opening_lot_headers(keys)
    out: list[dict[str, Any]] = []
    for values in rows_iter:
        row = {
            keys[idx]: values[idx] if idx < len(values) else None
            for idx in range(len(keys))
            if keys[idx] in OPENING_LOT_REQUIRED_KEYS
        }
        if any(_text(v) for v in row.values()):
            out.append(row)
    return out


def _load_opening_lot_rows(content: bytes, filename: str) -> list[dict[str, Any]]:
    lower = _text(filename).lower()
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return _read_opening_lot_xlsx(content)
    if lower.endswith(".csv") or lower.endswith(".txt"):
        return _read_opening_lot_csv(content)
    raise ValueError("Ondersteund formaat: .xlsx, .xlsm, .csv of .txt.")


def _normalize_opening_lot_rows(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for raw in rows:
        source_date = _parse_date(raw.get("source_date") or raw.get("date"))
        item = {
            "source_type": "opening_stock",
            "source_ref": "Opening LOT import",
            "supplier": _text(raw.get("supplier")),
            "lot_number": _text(raw.get("lot_number")),
            "sku_code": _text(raw.get("sku_code")),
            "product_name": _text(raw.get("product_name")),
            "source_date": source_date.isoformat() if source_date else "",
            "quantity": _num(raw.get("quantity")),
            "purchase_price_input": _num(raw.get("purchase_price_input")),
            "purchase_price_includes_excise": _bool(raw.get("purchase_price_includes_excise")),
            "excise_per_unit": _num(raw.get("excise_per_unit")),
            "packaging_cost_per_unit": 0.0,
            "other_direct_cost_per_unit": 0.0,
        }
        reasons: list[str] = []
        if not item["supplier"]:
            reasons.append("missing_supplier")
        if not item["lot_number"]:
            reasons.append("missing_lot")
        if not item["sku_code"]:
            reasons.append("missing_sku")
        if item["purchase_price_input"] <= 0:
            reasons.append("missing_purchase_price")
        item["status"] = "ok" if not reasons else "check"
        item["reasons"] = reasons
        items.append(item)
    summary = {
        "rows": len(items),
        "ok": sum(1 for item in items if item["status"] == "ok"),
        "check": sum(1 for item in items if item["status"] != "ok"),
        "missing_lot": sum(1 for item in items if "missing_lot" in item["reasons"]),
        "missing_sku": sum(1 for item in items if "missing_sku" in item["reasons"]),
        "missing_purchase_price": sum(1 for item in items if "missing_purchase_price" in item["reasons"]),
    }
    return items, summary


def preview_opening_lot_import(content: bytes, filename: str) -> dict[str, Any]:
    rows = _load_opening_lot_rows(content, filename)
    items, summary = _normalize_opening_lot_rows(rows)
    return {"summary": summary, "items": items[:500]}


def confirm_opening_lot_import(content: bytes, filename: str) -> dict[str, Any]:
    rows = _load_opening_lot_rows(content, filename)
    items, summary = _normalize_opening_lot_rows(rows)
    saved = 0
    for item in items:
        if item["status"] != "ok":
            continue
        upsert_lot_cost_record(item)
        saved += 1
    return {"summary": {**summary, "saved": saved}, "items": items[:500]}


def _chunk_pairs(pairs: list[tuple[str, str]], size: int = 250) -> list[list[tuple[str, str]]]:
    return [pairs[idx : idx + size] for idx in range(0, len(pairs), size)]


def _values_sql_for_pairs(pairs: list[tuple[str, str]]) -> tuple[str, list[str]]:
    placeholders = ",".join(["(%s, %s)"] * len(pairs))
    params: list[str] = []
    for tx, sku in pairs:
        params.extend([tx, sku])
    return placeholders, params


def _match_stock_rows(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    grouped: dict[tuple[str, str, str], dict[str, Any]] = {}
    for raw in rows:
        tx = _text(raw.get("transaction_number"))
        sku = _text(raw.get("sku_code"))
        lot = _text(raw.get("lot_number"))
        if not tx and not sku and not lot:
            continue
        key = (tx, sku, lot)
        item = grouped.setdefault(
            key,
            {
                "transaction_number": tx,
                "sku_code": sku,
                "lot_number": lot,
                "movement_date": _parse_date(raw.get("date")),
                "product_name": _text(raw.get("product_name")),
                "company_name": _text(raw.get("company_name")),
                "quantity": 0.0,
                "stock_value_per_unit": _num(raw.get("stock_value_per_unit")),
                "excise_per_unit": _num(raw.get("excise_per_unit")),
                "movement_type": _text(raw.get("movement_type")),
                "movement_reason": _text(raw.get("movement_reason")),
                "payload": {"examples": []},
            },
        )
        item["quantity"] = float(item.get("quantity", 0.0) or 0.0) + _num(raw.get("quantity"))
        examples = item["payload"].setdefault("examples", [])
        if len(examples) < 3:
            examples.append(raw)

    items = list(grouped.values())
    if not items:
        return [], {"rows": 0, "matched": 0, "unmatched": 0, "missing_lot": 0, "missing_sku": 0, "missing_transaction": 0}

    ensure_schema()
    postgres_storage.ensure_schema()
    from app.domain import douano_sync_storage

    douano_sync_storage.ensure_schema()
    pairs = sorted(
        {
            (_text(item.get("transaction_number")), _text(item.get("sku_code")))
            for item in items
            if _text(item.get("transaction_number")) and _text(item.get("sku_code"))
        }
    )
    invoice_matches: set[tuple[str, str]] = set()
    order_matches: set[tuple[str, str]] = set()
    if pairs:
        with postgres_storage.connect() as conn:
            with conn.cursor() as cur:
                for chunk in _chunk_pairs(pairs):
                    values_sql, params = _values_sql_for_pairs(chunk)
                    cur.execute(
                        f"""
                        WITH wanted(transaction_number, sku) AS (VALUES {values_sql})
                        SELECT DISTINCT w.transaction_number, w.sku
                        FROM wanted w
                        JOIN douano_sales_invoices i ON i.invoiced_transaction_numbers ? w.transaction_number
                        JOIN douano_sales_invoice_lines l ON l.sales_invoice_id = i.sales_invoice_id
                        JOIN douano_products p ON p.product_id = l.douano_product_id AND p.sku = w.sku
                        """,
                        tuple(params),
                    )
                    invoice_matches.update({(_text(tx), _text(sku)) for tx, sku in (cur.fetchall() or [])})
                    cur.execute(
                        f"""
                        WITH wanted(transaction_number, sku) AS (VALUES {values_sql})
                        SELECT DISTINCT w.transaction_number, w.sku
                        FROM wanted w
                        JOIN douano_sales_orders o ON o.transaction_number = w.transaction_number
                        JOIN douano_sales_order_lines l ON l.sales_order_id = o.sales_order_id
                        JOIN douano_products p ON p.product_id = l.douano_product_id AND p.sku = w.sku
                        """,
                        tuple(params),
                    )
                    order_matches.update({(_text(tx), _text(sku)) for tx, sku in (cur.fetchall() or [])})

    for item in items:
        tx = _text(item.get("transaction_number"))
        sku = _text(item.get("sku_code"))
        reasons: list[str] = []
        if not tx:
            reasons.append("missing_transaction")
        if not sku:
            reasons.append("missing_sku")
        if not _text(item.get("lot_number")):
            reasons.append("missing_lot")
        invoice_match = (tx, sku) in invoice_matches
        order_match = (tx, sku) in order_matches
        if tx and sku and not invoice_match and not order_match:
            reasons.append("no_douano_match")
        item["match"] = {
            "invoice": invoice_match,
            "order": order_match,
            "status": "matched" if invoice_match or order_match else "unmatched",
            "reasons": reasons,
        }

    summary = {
        "rows": len(items),
        "matched": sum(1 for item in items if item["match"]["status"] == "matched"),
        "unmatched": sum(1 for item in items if item["match"]["status"] != "matched"),
        "missing_lot": sum(1 for item in items if "missing_lot" in item["match"]["reasons"]),
        "missing_sku": sum(1 for item in items if "missing_sku" in item["match"]["reasons"]),
        "missing_transaction": sum(1 for item in items if "missing_transaction" in item["match"]["reasons"]),
    }
    return items, summary


def preview_stock_history_import(content: bytes, filename: str) -> dict[str, Any]:
    rows = _load_stock_history_rows(content, filename)
    items, summary = _match_stock_rows(rows)
    return {"summary": summary, "items": items[:500]}


def confirm_stock_history_import(content: bytes, filename: str) -> dict[str, Any]:
    rows = _load_stock_history_rows(content, filename)
    items, summary = _match_stock_rows(rows)
    ensure_schema()
    saved = 0
    now = _now()
    clean_filename = Path(_text(filename) or "voorraadhistoriek.xlsx").name
    batch_id = f"stock-history-{now.strftime('%Y%m%d%H%M%S')}-{uuid4().hex[:8]}"
    with postgres_storage.connect() as conn:
        with conn.cursor() as cur:
            for item in items:
                if not _text(item.get("transaction_number")) or not _text(item.get("sku_code")) or not _text(item.get("lot_number")):
                    continue
                rid = _id_for(batch_id, item.get("transaction_number"), item.get("sku_code"), item.get("lot_number"))
                cur.execute(
                    """
                    INSERT INTO sales_lot_allocations (
                        id, import_batch_id, source_filename, transaction_number, sku_code, lot_number, movement_date, product_name, company_name,
                        quantity, stock_value_per_unit, excise_per_unit, movement_type, movement_reason, payload, updated_at
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, %s)
                    ON CONFLICT (id) DO UPDATE SET
                        import_batch_id = EXCLUDED.import_batch_id,
                        source_filename = EXCLUDED.source_filename,
                        transaction_number = EXCLUDED.transaction_number,
                        sku_code = EXCLUDED.sku_code,
                        lot_number = EXCLUDED.lot_number,
                        movement_date = EXCLUDED.movement_date,
                        product_name = EXCLUDED.product_name,
                        company_name = EXCLUDED.company_name,
                        quantity = EXCLUDED.quantity,
                        stock_value_per_unit = EXCLUDED.stock_value_per_unit,
                        excise_per_unit = EXCLUDED.excise_per_unit,
                        movement_type = EXCLUDED.movement_type,
                        movement_reason = EXCLUDED.movement_reason,
                        payload = EXCLUDED.payload,
                        updated_at = EXCLUDED.updated_at
                    """,
                    (
                        rid,
                        batch_id,
                        clean_filename,
                        item["transaction_number"],
                        item["sku_code"],
                        item["lot_number"],
                        item.get("movement_date"),
                        item.get("product_name", ""),
                        item.get("company_name", ""),
                        float(item.get("quantity", 0.0) or 0.0),
                        float(item.get("stock_value_per_unit", 0.0) or 0.0),
                        float(item.get("excise_per_unit", 0.0) or 0.0),
                        item.get("movement_type", ""),
                        item.get("movement_reason", ""),
                        json.dumps(item, default=str),
                        now,
                    ),
                )
                saved += 1
        if not postgres_storage.in_transaction():
            conn.commit()
    return {"summary": {**summary, "saved": saved, "import_batch_id": batch_id}, "items": items[:500]}


def _sales_lot_row_id(*, source: str, transaction_number: Any, sku_code: Any, lot_number: Any, movement_date: Any, quantity: Any, stock_location: Any = "") -> str:
    return _id_for(source, transaction_number, sku_code, lot_number, movement_date, quantity, stock_location)


def upsert_douano_sales_lot_rows(rows: list[dict[str, Any]], *, source_ref: str = "douano_stock_history") -> dict[str, Any]:
    ensure_schema()
    now = _now()
    filtered: list[dict[str, Any]] = []
    skipped = 0
    for raw in rows:
        if _text(raw.get("transaction_type")).lower() != "verkoop":
            skipped += 1
            continue
        if _text(raw.get("stock_document_type")).lower() != "verzending":
            skipped += 1
            continue
        if _text(raw.get("cause")).lower() != "verwijderd":
            skipped += 1
            continue
        tx = _text(raw.get("transaction_number"))
        sku = _text(raw.get("sku"))
        if not tx or not sku:
            skipped += 1
            continue
        quantity = abs(_num(raw.get("quantity")))
        if quantity <= 0:
            skipped += 1
            continue
        filtered.append(
            {
                "id": _sales_lot_row_id(
                    source="douano",
                    transaction_number=tx,
                    sku_code=sku,
                    lot_number=_text(raw.get("lot_number")),
                    movement_date=_parse_date(raw.get("date")),
                    quantity=quantity,
                    stock_location=_text(raw.get("stock_location")),
                ),
                "transaction_number": tx,
                "sku_code": sku,
                "lot_number": _text(raw.get("lot_number")),
                "movement_date": _parse_date(raw.get("date")),
                "product_name": _text(raw.get("product")),
                "company_name": _text(raw.get("company")),
                "quantity": quantity,
                "stock_value_per_unit": _num(raw.get("value_per_unit")),
                "excise_per_unit": _num(raw.get("excise_per_unit")),
                "movement_type": _text(raw.get("transaction_type")),
                "movement_reason": _text(raw.get("cause")),
                "payload": {"source": "douano", "source_ref": source_ref, "raw": raw},
            }
        )

    saved = 0
    missing_lot = 0
    with postgres_storage.connect() as conn:
        with conn.cursor() as cur:
            for item in filtered:
                if not item["lot_number"]:
                    missing_lot += 1
                cur.execute(
                    """
                    INSERT INTO sales_lot_allocations (
                        id, import_batch_id, source_filename, transaction_number, sku_code, lot_number, movement_date, product_name, company_name,
                        quantity, stock_value_per_unit, excise_per_unit, movement_type, movement_reason, payload, updated_at
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, %s)
                    ON CONFLICT (id) DO UPDATE SET
                        transaction_number = EXCLUDED.transaction_number,
                        sku_code = EXCLUDED.sku_code,
                        lot_number = CASE
                            WHEN sales_lot_allocations.payload->>'lot_source' = 'excel_enrichment'
                              AND COALESCE(NULLIF(sales_lot_allocations.lot_number, ''), '') <> ''
                              AND COALESCE(NULLIF(EXCLUDED.lot_number, ''), '') = ''
                            THEN sales_lot_allocations.lot_number
                            ELSE EXCLUDED.lot_number
                        END,
                        movement_date = EXCLUDED.movement_date,
                        product_name = EXCLUDED.product_name,
                        company_name = EXCLUDED.company_name,
                        quantity = EXCLUDED.quantity,
                        stock_value_per_unit = EXCLUDED.stock_value_per_unit,
                        excise_per_unit = EXCLUDED.excise_per_unit,
                        movement_type = EXCLUDED.movement_type,
                        movement_reason = EXCLUDED.movement_reason,
                        payload = CASE
                            WHEN sales_lot_allocations.payload->>'lot_source' = 'excel_enrichment'
                              AND COALESCE(NULLIF(sales_lot_allocations.lot_number, ''), '') <> ''
                              AND COALESCE(NULLIF(EXCLUDED.lot_number, ''), '') = ''
                            THEN jsonb_set(EXCLUDED.payload, '{lot_source}', to_jsonb('excel_enrichment'::text), true)
                            ELSE EXCLUDED.payload
                        END,
                        updated_at = EXCLUDED.updated_at
                    """,
                    (
                        item["id"],
                        "douano_stock_history",
                        source_ref,
                        item["transaction_number"],
                        item["sku_code"],
                        item["lot_number"],
                        item["movement_date"],
                        item["product_name"],
                        item["company_name"],
                        item["quantity"],
                        item["stock_value_per_unit"],
                        item["excise_per_unit"],
                        item["movement_type"],
                        item["movement_reason"],
                        json.dumps(item["payload"], default=str),
                        now,
                    ),
                )
                saved += 1
        if not postgres_storage.in_transaction():
            conn.commit()
    return {"fetched": len(rows), "filtered": len(filtered), "saved": saved, "skipped": skipped, "missing_lot": missing_lot}


def enrich_missing_sales_lots_from_excel(content: bytes, filename: str) -> dict[str, Any]:
    rows = _load_stock_history_rows(content, filename)
    items, summary = _match_stock_rows(rows)
    updated = 0
    inserted = 0
    conflicts = 0
    missing_target = 0
    now = _now()
    clean_filename = Path(_text(filename) or "lot-verrijking.xlsx").name
    with postgres_storage.connect() as conn:
        with conn.cursor() as cur:
            for item in items:
                tx = _text(item.get("transaction_number"))
                sku = _text(item.get("sku_code"))
                lot = _text(item.get("lot_number"))
                if not tx or not sku or not lot:
                    continue
                cur.execute(
                    """
                    SELECT id, lot_number
                    FROM sales_lot_allocations
                    WHERE transaction_number = %s AND sku_code = %s
                    ORDER BY ABS(quantity) DESC, updated_at DESC
                    """,
                    (tx, sku),
                )
                candidates = cur.fetchall() or []
                target_id = ""
                conflict_found = False
                for rid, existing_lot in candidates:
                    existing = _text(existing_lot)
                    if not existing:
                        target_id = _text(rid)
                        break
                    if existing.lower() != lot.lower():
                        conflict_found = True
                if not target_id:
                    if conflict_found:
                        conflicts += 1
                    else:
                        match = item.get("match") if isinstance(item.get("match"), dict) else {}
                        if match.get("status") == "matched":
                            rid = _sales_lot_row_id(
                                source="excel_enrichment",
                                transaction_number=tx,
                                sku_code=sku,
                                lot_number=lot,
                                movement_date=item.get("movement_date"),
                                quantity=item.get("quantity"),
                                stock_location="",
                            )
                            cur.execute(
                                """
                                INSERT INTO sales_lot_allocations (
                                    id, import_batch_id, source_filename, transaction_number, sku_code, lot_number, movement_date, product_name, company_name,
                                    quantity, stock_value_per_unit, excise_per_unit, movement_type, movement_reason, payload, updated_at
                                )
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, %s)
                                ON CONFLICT (id) DO UPDATE SET
                                    lot_number = EXCLUDED.lot_number,
                                    movement_date = EXCLUDED.movement_date,
                                    product_name = EXCLUDED.product_name,
                                    company_name = EXCLUDED.company_name,
                                    quantity = EXCLUDED.quantity,
                                    stock_value_per_unit = EXCLUDED.stock_value_per_unit,
                                    excise_per_unit = EXCLUDED.excise_per_unit,
                                    movement_type = EXCLUDED.movement_type,
                                    movement_reason = EXCLUDED.movement_reason,
                                    payload = EXCLUDED.payload,
                                    updated_at = EXCLUDED.updated_at
                                """,
                                (
                                    rid,
                                    "excel_enrichment",
                                    clean_filename,
                                    tx,
                                    sku,
                                    lot,
                                    item.get("movement_date"),
                                    _text(item.get("product_name")),
                                    _text(item.get("company_name")),
                                    abs(_num(item.get("quantity"))),
                                    _num(item.get("stock_value_per_unit")),
                                    _num(item.get("excise_per_unit")),
                                    _text(item.get("movement_type")) or "Excel verrijking",
                                    _text(item.get("movement_reason")),
                                    json.dumps(
                                        {
                                            "source": "excel_enrichment",
                                            "source_ref": clean_filename,
                                            "lot_source": "excel_enrichment",
                                            "excel_enrichment_file": clean_filename,
                                            "match": match,
                                        },
                                        default=str,
                                    ),
                                    now,
                                ),
                            )
                            inserted += 1
                        else:
                            missing_target += 1
                    continue
                cur.execute(
                    """
                    UPDATE sales_lot_allocations
                    SET lot_number = %s,
                        source_filename = %s,
                        payload = jsonb_set(
                            jsonb_set(payload, '{lot_source}', to_jsonb('excel_enrichment'::text), true),
                            '{excel_enrichment_file}', to_jsonb(%s::text), true
                        ),
                        updated_at = %s
                    WHERE id = %s AND COALESCE(NULLIF(lot_number, ''), '') = ''
                    """,
                    (lot, clean_filename, clean_filename, now, target_id),
                )
                updated += int(cur.rowcount or 0)
        if not postgres_storage.in_transaction():
            conn.commit()
    return {
        "summary": {
            **summary,
            "updated": updated,
            "inserted": inserted,
            "conflicts": conflicts,
            "missing_target": missing_target,
        },
        "items": items[:500],
    }


def list_stock_history_imports(limit: int = 100) -> list[dict[str, Any]]:
    ensure_schema()
    lim = max(1, min(int(limit or 100), 1000))
    with postgres_storage.connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT import_batch_id, source_filename, COUNT(*) AS row_count, MAX(updated_at) AS imported_at
                FROM sales_lot_allocations
                WHERE import_batch_id <> ''
                GROUP BY import_batch_id, source_filename
                ORDER BY MAX(updated_at) DESC
                LIMIT %s
                """,
                (lim,),
            )
            rows = cur.fetchall() or []
    return [
        {
            "import_batch_id": _text(batch_id),
            "source_filename": _text(filename),
            "row_count": int(row_count or 0),
            "imported_at": imported_at.isoformat() if hasattr(imported_at, "isoformat") and imported_at else "",
        }
        for batch_id, filename, row_count, imported_at in rows
    ]


def delete_stock_history_import(import_batch_id: str) -> int:
    ensure_schema()
    batch = _text(import_batch_id)
    if not batch:
        raise ValueError("Import batch ontbreekt.")
    with postgres_storage.connect() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM sales_lot_allocations WHERE import_batch_id = %s", (batch,))
            deleted = int(cur.rowcount or 0)
        if not postgres_storage.in_transaction():
            conn.commit()
    return deleted
