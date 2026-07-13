from __future__ import annotations

import json
import sys
import urllib.error
import urllib.request
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


API_BASE = "http://localhost:8000/api"
LIST_PRICE_CODE = "list"
CHANNEL_COMPAT_CODES = [LIST_PRICE_CODE]


def api_json(path: str, method: str = "GET", payload: Any | None = None) -> Any:
    data = None if payload is None else json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        f"{API_BASE}{path}",
        data=data,
        method=method,
        headers={"Content-Type": "application/json"},
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as response:
            raw = response.read().decode("utf-8")
            return json.loads(raw) if raw else {}
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"{method} {path} failed: {exc.code} {body}") from exc


def read_price_list(path: Path, year: int) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Rapport 1"]
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    index = {name: idx for idx, name in enumerate(headers)}
    out: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = str(row[index["Artikelnummer"]] or "").strip()
        product = str(row[index["Product"]] or "").strip()
        price = row[index["Prijs"]]
        if not sku or price is None:
            continue
        try:
            parsed_price = round(float(price), 2)
        except (TypeError, ValueError):
            continue
        begin = row[index["Begindatum"]]
        end = row[index["Einddatum"]]
        out.append(
            {
                "year": year,
                "douano_sku": sku,
                "product": product,
                "price": parsed_price,
                "effective_from": begin.date().isoformat() if isinstance(begin, datetime) else str(begin or ""),
                "effective_to": end.date().isoformat() if isinstance(end, datetime) else str(end or ""),
            }
        )
    return out


def as_items(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, dict) and isinstance(value.get("items"), list):
        return [row for row in value["items"] if isinstance(row, dict)]
    if isinstance(value, list):
        return [row for row in value if isinstance(row, dict)]
    return []


def build_record(row: dict[str, Any], sku: dict[str, Any], douano_product_id: int) -> dict[str, Any]:
    kind = str(sku.get("kind", "") or "").strip().lower()
    sku_id = str(sku.get("id", "") or "").strip()
    beer_id = str(sku.get("beer_id", "") or "").strip()
    article_id = str(sku.get("article_id", "") or "").strip()
    format_id = str(sku.get("format_article_id", "") or "").strip()
    product_id = article_id if kind == "article" else format_id
    product_type = "samengesteld" if kind == "article" else "basis"
    prices = {code: row["price"] for code in CHANNEL_COMPAT_CODES}
    return {
        "id": f"verkoopstrategie-prijslijst-{row['year']}-{uuid.uuid5(uuid.NAMESPACE_URL, sku_id)}",
        "record_type": "verkoopstrategie_product",
        "jaar": row["year"],
        "bron_jaar": row["year"],
        "sku_id": sku_id,
        "douano_product_id": douano_product_id,
        "douano_sku": row["douano_sku"],
        "bier_id": beer_id,
        "biernaam": "",
        "product_id": product_id,
        "product_type": product_type,
        "verpakking": str(sku.get("name") or sku.get("naam") or row["product"] or sku_id),
        "strategie_type": "prijslijst_import",
        "price_list_name": f"Zakelijk {row['year']}",
        "price_status": "active",
        "effective_from": row["effective_from"],
        "effective_to": row["effective_to"],
        "kostprijs": 0,
        "kanaalmarges": {},
        "sell_in_margins": {},
        "kanaalprijzen": prices,
        "sell_in_prices": prices,
        "updated_from": f"Prijslijst{row['year']}.xlsx",
    }


def main() -> int:
    root = Path(__file__).resolve().parents[1]
    price_files = {
        2025: Path(r"C:\Users\hansh\Downloads\Prijslijst2025.xlsx"),
        2026: Path(r"C:\Users\hansh\Downloads\Prijslijst2026.xlsx"),
    }

    bootstrap = api_json(
        "/meta/bootstrap?datasets=verkoopprijzen%2Cskus%2Carticles%2Ckostprijsproductactiveringen%2Ckostprijsversies%2Cproductie%2Cbieren%2Cbasisproducten%2Csamengestelde-producten%2Cpackaging-component-prices%2Cbom-lines"
    )
    datasets = bootstrap.get("datasets", {}) if isinstance(bootstrap, dict) else {}
    skus = [row for row in datasets.get("skus", []) if isinstance(row, dict)]
    sku_by_id = {str(row.get("id", "") or ""): row for row in skus}

    products = as_items(api_json("/integrations/douano/products?limit=2000"))
    douano_by_sku = {str(row.get("sku", "") or "").strip(): row for row in products}
    mappings = as_items(api_json("/integrations/douano/product-mappings?limit=5000"))
    mapping_by_product_id = {int(row.get("douano_product_id", 0) or 0): row for row in mappings}

    current_items = as_items(api_json("/data/verkoopprijzen/items"))
    existing_by_year_sku: dict[tuple[int, str], dict[str, Any]] = {}
    for item in current_items:
        sku_id = str(item.get("sku_id", "") or "").strip()
        year = int(item.get("jaar", 0) or 0)
        if sku_id and year and str(item.get("record_type", "") or "") == "verkoopstrategie_product":
            existing_by_year_sku[(year, sku_id)] = item

    matched: list[dict[str, Any]] = []
    unmatched: list[dict[str, Any]] = []
    updated = 0
    created = 0
    next_items = list(current_items)
    excel_skus_by_year: dict[int, set[str]] = {}

    for year, path in price_files.items():
        excel_rows = read_price_list(path, year)
        excel_skus_by_year[year] = {row["douano_sku"] for row in excel_rows}
        for row in excel_rows:
            douano = douano_by_sku.get(row["douano_sku"])
            if not douano:
                unmatched.append({**row, "reason": "Douano product niet gevonden"})
                continue
            douano_product_id = int(douano.get("product_id", 0) or 0)
            mapping = mapping_by_product_id.get(douano_product_id)
            if not mapping:
                unmatched.append({**row, "reason": "Geen productkoppeling", "douano_product_id": douano_product_id})
                continue
            sku_id = str(mapping.get("sku_id", "") or "").strip()
            sku = sku_by_id.get(sku_id)
            if not sku:
                unmatched.append({**row, "reason": "Interne SKU niet gevonden", "douano_product_id": douano_product_id, "sku_id": sku_id})
                continue

            record = build_record(row, sku, douano_product_id)
            existing = existing_by_year_sku.get((year, sku_id))
            if existing:
                record["id"] = existing.get("id") or record["id"]
                merged = {**existing, **record}
                next_items = [merged if str(item.get("id", "")) == str(record["id"]) else item for item in next_items]
                updated += 1
            else:
                next_items.append(record)
                existing_by_year_sku[(year, sku_id)] = record
                created += 1
            matched.append(
                {
                    "year": year,
                    "douano_sku": row["douano_sku"],
                    "product": row["product"],
                    "price": row["price"],
                    "sku_id": sku_id,
                    "internal_name": sku.get("name") or sku.get("naam") or "",
                }
            )

    missing_app_skus: list[dict[str, Any]] = []
    for mapping in mappings:
        product_id = int(mapping.get("douano_product_id", 0) or 0)
        douano = next((row for row in products if int(row.get("product_id", 0) or 0) == product_id), None)
        if not douano:
            continue
        douano_sku = str(douano.get("sku", "") or "").strip()
        sku_id = str(mapping.get("sku_id", "") or "").strip()
        sku = sku_by_id.get(sku_id)
        if not sku:
            continue
        for year in price_files:
            if douano_sku not in excel_skus_by_year.get(year, set()):
                missing_app_skus.append(
                    {
                        "year": year,
                        "douano_sku": douano_sku,
                        "douano_product": douano.get("name"),
                        "sku_id": sku_id,
                        "internal_name": sku.get("name") or sku.get("naam") or "",
                    }
                )

    api_json("/data/verkoopprijzen", method="PUT", payload=next_items)

    report = {
        "created": created,
        "updated": updated,
        "matched_count": len(matched),
        "unmatched_count": len(unmatched),
        "missing_app_skus_count": len(missing_app_skus),
        "matched": matched,
        "unmatched": unmatched,
        "missing_app_skus": missing_app_skus,
    }
    report_path = root / "outputs" / "price-list-import-report.json"
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps({k: report[k] for k in ["created", "updated", "matched_count", "unmatched_count", "missing_app_skus_count"]}, indent=2))
    print(str(report_path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
